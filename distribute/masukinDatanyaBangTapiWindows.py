import pandas as pd
from openpyxl import load_workbook

def merge_data(input_files, output_file, template_file):
    template = load_workbook(template_file)

    for input_file in input_files:
        name = input_file.split("_")[0]

        # Mengubah nama sheet menjadi lowercase
        sheet_name = name.lower()

        # Mengecek jika nama sheet sudah ada, jika tidak ada, tambahkan sheet baru
        if sheet_name not in template.sheetnames:
            template.create_sheet(sheet_name)

        sheet = template[sheet_name]

        df = pd.read_excel(input_file, header=None)

        data_range = df.loc[:249, :9].values.tolist()

        start_cell = sheet['B10']

        for r, row in enumerate(data_range):
            for c, value in enumerate(row):
                sheet.cell(row=start_cell.row + r, column=start_cell.column + c).value = value

    template.save(output_file)
    print(f"Merged data saved to: {output_file}")

# Penggunaan
input_files = [f"{i}_1.xlsx" for i in range(1, 20)]


output_file = "MATENGS.xlsx"
template_file = "./Template/templateabsen.xlsx"

merge_data(input_files, output_file, template_file)
