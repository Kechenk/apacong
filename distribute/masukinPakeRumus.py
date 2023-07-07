import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def update_master_sheet():
    # Load the MASTER sheet
    wb = load_workbook('Absensi Komunitas 07-08.xlsx')
    master_sheet = wb['MASTER']

    # Iterate through the sheets and update the MASTER sheet formulas
    for sheet_number in range(1, 37):
        sheet_name = str(sheet_number)
        if sheet_name != 'MASTER':
            start_column = 2
            for col in range(start_column, start_column + 14):  # Columns B to N
                formula_top = f"='{sheet_name}'!{get_column_letter(col)}10"
                row_top = 10 + (sheet_number - 1) * 250

                cell_top = master_sheet.cell(row=row_top, column=col)
                cell_top.value = formula_top

                for i in range(1, 250):  # Fill data from B11 to B259
                    formula = f"='{sheet_name}'!{get_column_letter(col)}{10 + i}"
                    row = row_top + i

                    cell = master_sheet.cell(row=row, column=col)
                    cell.value = formula

    # Save the updated workbook
    wb.save('Absensi Komunitas 07-08.xlsx')
    print("MASTER sheet formulas updated.")

# Usage
update_master_sheet()
