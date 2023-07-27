import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection

def copy_cell_style(source_cell, target_cell):
    target_cell.font = Font(size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            color=source_cell.font.color)

    target_cell.border = Border(left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom,
                                diagonal=source_cell.border.diagonal,
                                diagonal_direction=source_cell.border.diagonal_direction,
                                outline=source_cell.border.outline,
                                vertical=source_cell.border.vertical,
                                horizontal=source_cell.border.horizontal)

    target_cell.fill = PatternFill(start_color=source_cell.fill.start_color,
                                   end_color=source_cell.fill.end_color,
                                   fill_type=source_cell.fill.fill_type)

    target_cell.number_format = source_cell.number_format

    # Handle protection style
    target_cell.protection = Protection(locked=source_cell.protection.locked,
                                        hidden=source_cell.protection.hidden)

    # Handle alignment style
    target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                      vertical=source_cell.alignment.vertical,
                                      text_rotation=source_cell.alignment.text_rotation,
                                      wrap_text=source_cell.alignment.wrap_text,
                                      shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                      indent=source_cell.alignment.indent,
                                      relative_indent=source_cell.alignment.relative_indent,
                                      justify_last_line=source_cell.alignment.justify_last_line,
                                      reading_order=source_cell.alignment.reading_order)

def main(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    master_sheet = wb.create_sheet("MASTER") 

    for sheet_name in wb.sheetnames:
        if sheet_name == "MASTER":
            continue 

        sheet = wb[sheet_name]

        for row in range(10, 260):
            value = sheet.cell(row=row, column=2).value
            fill = sheet.cell(row=row, column=2).fill

            master_cell = master_sheet.cell(row=row, column=2)
            master_cell.value = value
            copy_cell_style(sheet.cell(row=row, column=2), master_cell)

    wb.save(output_file)

input_file = 'markicobs.xlsx'
output_file = 'udah.xlsx'
main(input_file, output_file)
