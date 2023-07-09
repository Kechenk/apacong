import os.path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

scopes = ["https://www.googleapis.com/auth/spreadsheets"]
spreadsheet_id = "1eXDnzC-yarKhS-n4SPpaqMSNU3C7FVtU"

#def # update_master_sheet():
    ## Load the MASTER sheet
    #wb = load_workbook('test data.xlsx')
    #master_sheet = wb['MASTER']

    # Iterate through the sheets and update the MASTER sheet formulas
    for sheet_number in range(1, 37):
        sheet_name = str(sheet_number)
        if sheet_name != 'MASTER':
            start_column = 2
            for col in range(start_column, start_column + 14):  # Columns B to N
                formula_top = f"='{sheet_name}'!{get_column_letter(col)}10"
                row_top = 10 + (sheet_number - 1) * 250

                cell_top = master_sheet.cell(row=row_top, column=col)
                cell_top.value = formula_top

                for i in range(1, 250):  # Fill data from B11 to B259
                    formula = f"='{sheet_name}'!{get_column_letter(col)}{10 + i}"
                    row = row_top + i

                    cell = master_sheet.cell(row=row, column=col)
                    cell.value = formula

    # Save the updated workbook
    wb.save('test data.xlsx')
    print("MASTER sheet formulas updated.")

def main():
    credentials = None
    if os.path.exists("token.json"):
        credentials = Credentials.from_authorized_user_file("token.json", scopes)
    if not credentials or not credentials.valid:
        if credentials and credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", scopes)
            credentials = flow.run_local_server(port=0)
        with open("token.json", "w") as token:
            token.write(credentials.to_json())

    service = build("sheets", "v4", credentials=credentials)
    sheet = service.spreadsheets()

    # Update the MASTER sheet
    # update_master_sheet()

    # Perform other spreadsheet operations using the 'sheet' object

if __name__ == "__main__":
    main()
