import openpyxl

def main(input_file, output_file):
	wb = openpyxl.load_workbook(input_file)
	master = wb.create_sheet("MASTER")

	for sheet_name in wb.sheetnames:
		if sheet_name == "MASTER":
			continue

		sheet = wb[sheet_name]

		for row in range(10, 260):

			value = sheet.cell(row=row, column=2).value

			fill = sheet.cell(row=row, column=2).fill

			master_cell = master_sheet.cell(row=row, column=2)
			master_cell.value = value
			master_cell.fill = fill

		wb.save(output_file)

input_file = 'markicobs.xlsx'
output_file = 'udah.xlsx'
main(input_file, output_file)