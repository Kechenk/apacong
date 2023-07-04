import pandas as pd
from openpyxl import load_workbook

def merge_data(input_files, output_file, template_file):
    # Load the template fil
    template = load_workbook(template_file)

    for input_file in input_files:
        # Extract the name from the input file name
        name = input_file.split("_")[0]

        # Get the corresponding sheet in the template file
        sheet = template[name]

        # Load the input file data
        df = pd.read_excel(input_file, header=None)

        # Get the data range in the input file
        data_range = df.loc[:249, :7].values.tolist()

        # Get the starting cell in the template sheet based on the sheet name
        start_cell = template[name]['C10']

        # Copy the data to the template sheet
        for r, row in enumerate(data_range):
            for c, value in enumerate(row):
                sheet.cell(row=start_cell.row + r, column=start_cell.column + c).value = value

    # Save the modified template file
    template.save(output_file)
    print(f"Merged data saved to: {output_file}")

# Usage
input_files = [
"1_1.xlsx", "1_2.xlsx", "2_1.xlsx", "2_2.xlsx", "3_1.xlsx", "3_2.xlsx", "4_1.xlsx", "4_2.xlsx", "5_1.xlsx", "5_2.xlsx",
"6_1.xlsx", "6_2.xlsx", "7_1.xlsx", "7_2.xlsx", "8_1.xlsx", "8_2.xlsx", "9_1.xlsx", "9_2.xlsx", "10_1.xlsx", "10_2.xlsx",
"11_1.xlsx", "11_2.xlsx", "12_1.xlsx", "12_2.xlsx", "13_1.xlsx", "13_2.xlsx", "14_1.xlsx", "14_2.xlsx", "15_1.xlsx",
"15_2.xlsx", "16_1.xlsx", "16_2.xlsx", "17_1.xlsx", "17_2.xlsx", "18_1.xlsx", "18_2.xlsx", "19_1.xlsx"
]

output_file = "merged_data.xlsx"
template_file = "templateabsen.xlsx"

merge_data(input_files, output_file, template_file)
