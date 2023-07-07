import pandas as pd
from openpyxl import load_workbook

def merge_data(input_file, output_file):
    template = load_workbook(input_file)
    master_sheet = template.active

    start_row = 10  # Starting row for data in the master sheet

    for sheet_number in range(1, 37):
        sheet_name = str(sheet_number)
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)

        # Copy the data from the input sheet to the master sheet
        data_range = df.loc[9:258, 1:13].values.tolist()
        for r, row in enumerate(data_range):
            for c, value in enumerate(row):
                master_sheet.cell(row=start_row + r, column=c + 2).value = value

        start_row += len(data_range)  # Increment the starting row for the next sheet

    template.save(output_file)
    print(f"Merged data saved to: {output_file}")

# Usage
input_file = "Absensi Komunitas 07-08.xlsx"
output_file = "Merged_Data.xlsx"

merge_data(input_file, output_file)
