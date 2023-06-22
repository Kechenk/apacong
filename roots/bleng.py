import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('modified_chunk.xlsx')

# Select the desired worksheet
worksheet = workbook['Sheet1']  # Ganti 'Sheet1' dengan nama lembar kerjamu

# Loop melalui setiap baris dari bawah ke atas
max_row = worksheet.max_row
for row in range(10, max_row + 1):
    is_empty = True
    for column in range(3, 11):
        cell = worksheet.cell(row=row, column=column)
        if cell.value is not None:
            is_empty = False
            break
    if is_empty:
        worksheet.delete_rows(row)

# Simpan file Excel yang sudah dimodifikasi
workbook.save('no200k.xlsx')
